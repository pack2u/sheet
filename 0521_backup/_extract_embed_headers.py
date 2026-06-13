# -*- coding: utf-8 -*-
"""Extract vendor headers from 업체별대리발송.xlsx → JSON for EMBEDDED_VENDOR_EXCLUSIVE_MASTER_ROWS_."""
import pathlib
import json
import re

import openpyxl

root = pathlib.Path(__file__).resolve().parent
f = root / "업체별대리발송.xlsx"

SKIP_SHEETS = {
    "주문-양식O",
    "판매중",
    "시트42",
    "업체미분류",
    "발송불가",
    "금일주문매핑추가",
    "누적품목매핑",
    "양식변환발주처목록",
    "주문-상품명변환O",
    "업체추가순서",
    "뉴파츠단가",
    "BOM현황",
    "도매문의",
    "화물차별적재량",
    "import대리발송",
    "(삭제대기)누적품목매핑",
}

# 업체명(목록) → 실제 시트 탭 이름
VENDOR_TAB_ALIASES = {
    "올팩코리아": "올팩",
    "제이씨인터내셔널": "제이씨",
    "뉴파츠": "뉴파츠_NEW",
}


def norm_cell(v):
    if v is None:
        return ""
    s = str(v).replace("\r\n", "\n").replace("\r", "\n")
    s = re.sub(r"\s+", " ", s).strip()
    return s


def row_to_pipe_csv(raw_row):
    parts = [norm_cell(x) for x in raw_row]
    while parts and parts[-1] == "":
        parts.pop()
    parts = [p for p in parts if p]
    return "|".join(parts)


def sheet_first_nonempty_row(ws, max_scan=40):
    mc = min(ws.max_column or 1, 80)
    for r in range(1, max_scan + 1):
        vals = [ws.cell(r, c).value for c in range(1, mc + 1)]
        if any(v is not None and norm_cell(v) for v in vals):
            return vals
    return []


def resolve_sheet_tab(vendor_name, sheetnames):
    vn = norm_cell(vendor_name)
    if not vn:
        return None
    if vn in sheetnames:
        return vn
    if vn in VENDOR_TAB_ALIASES:
        alias = VENDOR_TAB_ALIASES[vn]
        if alias in sheetnames:
            return alias
    for sn in sheetnames:
        if sn in SKIP_SHEETS:
            continue
        if vn == sn or vn in sn or sn in vn:
            return sn
    return None


def main():
    wb = openpyxl.load_workbook(f, read_only=False, data_only=True)
    sheetnames = set(wb.sheetnames)

    vendors = []
    if "양식변환발주처목록" in wb.sheetnames:
        sh = wb["양식변환발주처목록"]
        for row in sh.iter_rows(min_row=2, max_row=500, values_only=True):
            if not row or len(row) < 2:
                continue
            code = norm_cell(row[0])
            vname = norm_cell(row[1])
            if not vname:
                continue
            tab = resolve_sheet_tab(vname, wb.sheetnames)
            if not tab:
                continue
            vendors.append({"prefix": code.upper(), "tab": tab})

    seen = set()
    out = []
    for v in vendors:
        tab = v["tab"]
        if tab in seen:
            continue
        seen.add(tab)
        ws = wb[tab]
        vals = sheet_first_nonempty_row(ws)
        csv_line = row_to_pipe_csv(vals)
        if not csv_line:
            continue
        out.append(
            {
                "label": tab,
                "prefix": v["prefix"] or "",
                "headerCsv": csv_line,
            }
        )

    wb.close()

    lines = []
    lines.append("/**")
    lines.append(
        " * 업체별대리발송.xlsx 의 「양식변환발주처목록」+ 각 업체 탭 1행 기준 자동 생성."
    )
    lines.append(
        " * 엑셀 갱신 후 동일 폴더에서 `python _extract_embed_headers.py` 로 재생성."
    )
    lines.append(" */")
    lines.append(
        "var EMBEDDED_VENDOR_EXCLUSIVE_MASTER_ROWS_ = "
        + json.dumps(out, ensure_ascii=False, indent=2)
        + ";"
    )
    text = "\n".join(lines) + "\n"
    out_path = root / "_embedded_vendor_exclusive.generated.txt"
    out_path.write_text(text, encoding="utf-8")
    print("Wrote", out_path)


if __name__ == "__main__":
    main()
