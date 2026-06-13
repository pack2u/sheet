# -*- coding: utf-8 -*-
"""업체별대리발송.xlsx의 각 업체 탭 수식 구조를 분석."""
import pathlib
import openpyxl

root = pathlib.Path(__file__).resolve().parent
f = root / "업체별대리발송.xlsx"

SKIP_SHEETS = {
    "주문-양식O", "판매중", "시트42", "업체미분류", "발송불가",
    "금일주문매핑추가", "누적품목매핑", "양식변환발주처목록",
    "주문-상품명변환O", "업체추가순서", "뉴파츠단가", "BOM현황",
    "도매문의", "화물차별적재량", "import대리발송",
    "(삭제대기)누적품목매핑", "대리발송",
}

wb = openpyxl.load_workbook(f, read_only=False, data_only=False)
print(f"=== 전체 시트 목록 ({len(wb.sheetnames)}개) ===")
for name in wb.sheetnames:
    skip = "(SKIP)" if name in SKIP_SHEETS else ""
    print(f"  - {name} {skip}")

print("\n" + "="*70)

for name in wb.sheetnames:
    if name in SKIP_SHEETS:
        continue
    ws = wb[name]
    mc = min(ws.max_column or 1, 32)
    mr = min(ws.max_row or 1, 5)

    print(f"\n=== [{name}] ({ws.max_row}행 x {ws.max_column}열) ===")

    for r in range(1, mr + 1):
        print(f"  --- {r}행 ---")
        for c in range(1, mc + 1):
            cell = ws.cell(r, c)
            col_letter = openpyxl.utils.get_column_letter(c)
            val = cell.value
            if val is None:
                continue
            s = str(val)
            if s.startswith("="):
                print(f"    {col_letter}{r}: [수식] {s}")
            else:
                # 값이 길면 축약
                disp = s[:60] + "..." if len(s) > 60 else s
                print(f"    {col_letter}{r}: {disp}")

wb.close()
